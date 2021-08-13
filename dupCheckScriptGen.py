#! python3

# dupCheckScriptGen.py

# Use datastructures like a dictionary to load data that can be used for a dup checker
# use the snowflake connector to load this data and generate and execute correct group-by 
# dup checker scripts

import snowflake.connector as sf
# had to pip3 install --user snowflake-connector-python[pandas]
# and then path to the location (I don't use anaconda).. part of below doc helped
# https://navhealth.atlassian.net/wiki/spaces/RES/pages/2224095370/Setup+Python3+and+Anaconda+on+Mac

import pandas as pd
from tabulate import tabulate
import pprint
import pyinputplus as pyip
from openpyxl import load_workbook
import time

tableKeysStruct = {}

strtTime = time.time()

wb = load_workbook('/Users/michael.oconnor/sfIDFile.xlsx')
ws = wb.active
usrID = ws['A2'].value
pssWrd = ws['B2'].value
wb.close()


tableType = ''
orgDB = ''

# tableType = pyip.inputMenu(['Not Network Not Layup','Network Not Layup','Layup'], lettered=False, numbered=True)
tableType = pyip.inputMenu(['Base','Network','Layup'], lettered=False, numbered=True)

orgDB = pyip.inputMenu(['PROD_BEAUMONT', 'PROD_CHENMED', 'PROD_CLOVERNA', 'PROD_EVOLENT', 'PROD_OAKSTREETNA', 'PROD_PARADIGM', 'PROD_PRIVIA', 'PROD_VILLAGEMDNA'], lettered=False, numbered=True)

# 'PROD_BRIGHT', 'PROD_OSCAR', 
# Bright and Oscar are old orgs...Oscar and Bright both have time stamp in June 2021.   Privia has July 2021.
# I think PROFILE_LIST_OUTPATIENT_NETWORK_LAYUP is a known dup but idk about RNDRG_PROVIDER_SAVINGS_OPP_NETWORK_LAYUP
# DC_PROVIDER_BENCHMARK_NETWORK_LAYUP many dups for Privia

try:
    sf.paramstyle = 'qmark'
    sf_conn = sf.connect(
        # user=user, # Snowflake user e.g. first_name.last_name@carejourney.com
        user=usrID,
        # password=password, # carejourney-prod Okta password
        password=pssWrd,
        account='carejourney_nci.us-east-1',
        authenticator='externalbrowser',
        warehouse='DEV_BRIGHT', # snowflake warehouse --optional
        database='DEV_BRIGHT' # snowflake db --optional
    )

    sf_cursor = sf_conn.cursor()

    sql = ''
    tableWoutPK = []

    # need access to data_model db in nci  ... datamodel sometimes gets ahead (out of synch) with orgDB
    # if tableType == 'Not Network Not Layup': #? base 48
    #     sql = 'SELECT table_name FROM data_model.information_schema.TABLES WHERE table_schema LIKE \'VRDC%\' AND TABLE_name NOT LIKE \'%NETWORK%\' AND table_name NOT LIKE \'%LAYUP%\' order by table_name'
    # elif tableType == 'Network Not Layup': # 43
    #     sql = 'SELECT table_name FROM data_model.information_schema.TABLES WHERE table_schema LIKE \'VRDC%\' AND TABLE_name LIKE \'%NETWORK%\' AND table_name NOT LIKE \'%LAYUP%\' order by table_name'    
    # else:  # layup 24 includes network_layup
    #     sql = 'SELECT table_name FROM data_model.information_schema.TABLES WHERE table_schema LIKE \'VRDC%\' AND TABLE_name LIKE \'%LAYUP%\' order by table_name'
    # intended on using data_model.information_schema but it can get out of synch with
    # orgdb's information_schema... can get out of synch the other way too where org's db has a table that
    # datamodel doesn't.. need to inner join
    if tableType == 'Base': #? base 48
        # sql = 'SELECT table_name FROM data_model.information_schema.TABLES WHERE table_schema LIKE \'VRDC%\' AND TABLE_name NOT LIKE \'%NETWORK%\' AND table_name NOT LIKE \'%LAYUP%\' order by table_name'
        sql = '''SELECT orgDBT.table_name 
                FROM {orgDB}.information_schema.TABLES orgDBT
                join data_model.information_schema.tables dbT
                    on orgDBT.table_name = dbT.table_name
                WHERE orgDBT.table_schema LIKE 'VRDC%' AND orgDBT.TABLE_name NOT LIKE '%NETWORK%' 
                AND orgDBT.table_name NOT LIKE '%LAYUP%' 
                order by orgDBT.table_name'''.format(orgDB=orgDB)
    elif tableType == 'Network': # 43
        # sql = 'SELECT table_name FROM data_model.information_schema.TABLES WHERE table_schema LIKE \'VRDC%\' AND TABLE_name LIKE \'%NETWORK%\' AND table_name NOT LIKE \'%LAYUP%\' order by table_name'
        # FE doesn't have PPI_NETWORK TABLES yet
        sql = '''SELECT orgDBT.table_name 
                FROM {orgDB}.information_schema.TABLES orgDBT
                join data_model.information_schema.tables dbT
                    on orgDBT.table_name = dbT.table_name                
                WHERE orgDBT.table_schema LIKE 'VRDC%' AND orgDBT.TABLE_name LIKE '%NETWORK%' 
                AND orgDBT.table_name NOT LIKE '%LAYUP%' 
                AND orgDBT.TABLE_name NOT LIKE 'PPI_%' 
                order by orgDBT.table_name'''.format(orgDB=orgDB)
    else:  # layup 24 includes network_layup
        # sql = 'SELECT table_name FROM data_model.information_schema.TABLES WHERE table_schema LIKE \'VRDC%\' AND TABLE_name LIKE \'%LAYUP%\' order by table_name'
        sql = '''SELECT orgDBT.table_name 
                FROM {orgDB}.information_schema.TABLES orgDBT
                join data_model.information_schema.tables dbT
                    on orgDBT.table_name = dbT.table_name                 
                WHERE orgDBT.table_schema LIKE 'VRDC%' 
                AND orgDBT.TABLE_name LIKE '%LAYUP%' 
                order by orgDBT.table_name'''.format(orgDB=orgDB)


    sf_cursor.execute(sql)
    df = sf_cursor.fetch_pandas_all()
    singleColTN=df.loc[:,'TABLE_NAME']

    # a list of all tables in the schema, there's 115 in vrdc in total
    tblList = singleColTN.tolist()

    # tblList = ['PROFILE_LIST_RNDRG_PG','PROFILE_LIST_RNDRG_PG_NETWORK']

    # test with the first 15 tables
    for i in range(len(tblList)):
        # need access to data_model.vrdc schema
        sqlString = 'SHOW PRIMARY KEYS IN TABLE DATA_MODEL.VRDC.' + tblList[i]
        sf_cursor.execute(sqlString)
        table_results = []
        results = sf_cursor.fetchall()  # df = sf_cursor.fetch_pandas_all()  throws an 'unknown' error ? because no column names? 

        # print(str(len(results)))
        for idx in range(len(results)):
            table_results.append([results[idx][3], results[idx][4]])

        # print(tabulate(table_results, headers=["table_name", "pk_name"], tablefmt='psql'))
        df2 = pd.DataFrame(table_results, columns=["table_name", "pk_name"])
        singleColPK = df2.loc[:,"pk_name"]
        tblPKList = singleColPK.tolist()
        # can I create the pk list without pandas above?

        # load the dictionary
        if len(tblPKList) > 0:
            tableKeysStruct[tblList[i]] = tblPKList
        else:
            # print(tblList[i] + ' has no keys')
            tableWoutPK.append(tblList[i])

    # pprint.pprint(tableKeysStruct)

    # tableKeys = tableKeysStruct.keys()
    # df3 = pd.DataFrame(tableKeys, columns=["table_name"])
        # overwrites if exists
    # df3.to_excel('/Users/michael.oconnor/downloads/myTables3.xlsx',sheet_name=tableType)

    def dupCheckTest(tableDict):
        print('running dup check...')
        table_results = []
        for table, pks in tableDict.items():
            # print(table, ', '.join(tableDict[table]))
            # print(table, ', '.join(pks))
            sqlString = '''WITH dupCTE
                            AS
                            (
                            SELECT '{table}' AS tableName
                                , sum(rowsDuped) AS rowsAffected
                                , count(*) AS pkRowCount
                            FROM
                                (
                                SELECT {compositeKeys}
                                    , count(*) AS rowsDuped 
                                FROM {orgDB}.VRDC.{table}
                                GROUP BY {compositeKeys}
                                HAVING count(*) > 1
                                ) a 
                            ) 
                            SELECT tableName 
                                , '{orgDB}' AS orgDBName
                                , rowsAffected
                                , pkRowCount
                            FROM dupCTE'''.format(table=table,orgDB=orgDB,compositeKeys=', '.join(pks))                            
            # print(sqlString)
            sf_cursor.execute(sqlString)
            results = sf_cursor.fetchall()
            table_results.append([results[0][0], results[0][1], results[0][2], results[0][3]])

        df5 = pd.DataFrame(table_results, columns=["tableName", "orgDBName", "rowsAffected", "pkRowCount"]) 
        return df5   

    # need to cast load_ts to date here too

    def test2(tableDict):
        print('running test2. prod row counts..')
        table_results = []
        for table in tableDict.keys():
            sqlString = '''SELECT '{table}' AS TableName
                            , '{orgDB}' AS ordDBName
                            , max(date(load_ts)) AS MaxLoadTS
                            , max(load_period) AS MaxLP
                            , count(*) AS rwCount
                        FROM {orgDB}.VRDC.{table}'''.format(table=table,orgDB=orgDB)
            sf_cursor.execute(sqlString)
            results = sf_cursor.fetchall()
            # df3 = sf_cursor.fetch_pandas_all() # overwrites each time. need to append to a list like above
            # print(df3)
            # df4 += df3
            table_results.append([results[0][0], results[0][1], results[0][2], results[0][3], results[0][4]])

        df3 = pd.DataFrame(table_results, columns=["tableName", "orgDBName", "MaxLoadTS", "MaxLP", "rwCount"])
        return df3

    
    

    orgDB_FE =  orgDB + '_FE'

    # need to cast load_ts to date here
    # does the call to test2 and the return to df_prod need to be here? 

    def test3(tableDict):
        print('running test3..prod_fe row counts..')
        table_results = []
        for table in tableDict.keys():
            sqlString = '''SELECT '{table}' AS TableName
                            , '{orgDB}' AS ordDBName
                            , max(date(load_ts)) AS MaxLoadTS
                            , max(load_period) AS MaxLP                            
                            , count(*) AS rwCount
                        FROM {orgDB}.VRDC.{table}'''.format(table=table,orgDB=orgDB_FE)
            sf_cursor.execute(sqlString)
            results = sf_cursor.fetchall()
            # df3 = sf_cursor.fetch_pandas_all() # overwrites each time. need to append to a list like above
            # print(df3)
            # df4 += df3
            table_results.append([results[0][0], results[0][1], results[0][2], results[0][3], results[0][4]])

        df3 = pd.DataFrame(table_results, columns=["tableName", "orgDBName", "MaxLoadTS", "MaxLP", "rwCount"])
        # df3['MaxLoadTS'] = df3['MaxLoadTS'].dt.date   # hack  Can only use .dt accessor with datetimelike value
        df4 = pd.merge(df_prod, df3, how='left', on='tableName')
        df4['diff'] = df4['rwCount_x'] - df4['rwCount_y'] # does this work?  
        return df4

    def test4(tableDict):
        print('running test 4... utilization in Prod')
        for table in tableDict.keys():
            if table == 'PROFILE_LIST_RNDRG_PG' or table == 'PROFILE_LIST_RNDRG_PG_NETWORK':
                sqlString = '''with pac_count as (
                                    select 
                                        YEAR
                                        , 'PAC Count' as label
                                        , count(distinct PAC_NUM) as METRIC_TOTAL
                                        from {orgDB}.VRDC.{table}
                                        group by 1
                                    )
                                    , metrics as (
                                    select 
                                        YEAR
                                        , METRIC_LABEL as LABEL
                                        , SUM(METRIC) as METRIC_TOTAL
                                        from {orgDB}.VRDC.{table}
                                        where METRIC_LABEL in ('Total Allowed - Inpatient Hospital',
                                                        'Total Allowed - Outpatient Hospital',
                                                        'Total Allowed - Office',
                                                        'Total Allowed - Hospice',
                                                        'Total Allowed - Home', 
                                                        'Total Allowed - Pharmacy'
                                                        )
                                        group by 1, 2
                                    )
                                    select '{table}' as tableName, year, label, '{orgDB}' as orgDB, metric_total from pac_count
                                    union all
                                    select '{table}' as tableName, year, label, '{orgDB}' as orgDB, metric_total from metrics
                                    order by 1, 2, 3
                                    ;'''.format(table=table,orgDB=orgDB)

                sf_cursor.execute(sqlString)
                # how do you append outputs/stack them for pg and pg_network if both can have several rows?   
                df = sf_cursor.fetch_pandas_all()
                # df += df.append(df)
        return df

    def test5(tableDict):
        print('running test 5... utilization in FE')
        for table in tableDict.keys():
            if table == 'PROFILE_LIST_RNDRG_PG' or table == 'PROFILE_LIST_RNDRG_PG_NETWORK':
                sqlString = '''with pac_count as (
                                    select 
                                        YEAR
                                        , 'PAC Count' as label
                                        , count(distinct PAC_NUM) as METRIC_TOTAL
                                        from {orgDB_FE}.VRDC.{table}
                                        group by 1
                                    )
                                    , metrics as (
                                    select 
                                        YEAR
                                        , METRIC_LABEL as LABEL
                                        , SUM(METRIC) as METRIC_TOTAL
                                        from {orgDB_FE}.VRDC.{table}
                                        where METRIC_LABEL in ('Total Allowed - Inpatient Hospital',
                                                        'Total Allowed - Outpatient Hospital',
                                                        'Total Allowed - Office',
                                                        'Total Allowed - Hospice',
                                                        'Total Allowed - Home', 
                                                        'Total Allowed - Pharmacy'
                                                        )
                                        group by 1, 2
                                    )
                                    select '{table}' as tableName, year, label, '{orgDB_FE}' as orgDB, metric_total from pac_count
                                    union all
                                    select '{table}' as tableName, year, label, '{orgDB_FE}' as orgDB, metric_total from metrics
                                    order by 1, 2, 3
                                    ;'''.format(table=table,orgDB_FE=orgDB_FE)

                sf_cursor.execute(sqlString)
                # how do you append outputs for pg and pg_network if both can have several rows?   
                df = sf_cursor.fetch_pandas_all()
                # print(df)
                # print(dfPG)
                df5 = pd.merge(dfPG, df, how='left', on=['TABLENAME','YEAR','LABEL'])
                df5['diff'] = df5['METRIC_TOTAL_x'] - df5['METRIC_TOTAL_y'] 
                #['tableName','year','label']
                # print(df5)
        return df5

    def test6(tableDict):
        print('running test 6..cji utilization in PROD..')
        for table in tableDict.keys():
            if table == 'CJI_COST_SCORECARD' or table == 'CJI_COST_SCORECARD_NETWORK':
                sqlString = '''with npi_count as (
                                       select 
                                          YEAR
                                        , 'NPI Count' as label
                                        , count(distinct NPI) as metric_total
                                        from {orgDB}.VRDC.{table}
                                        group by 1
                                    )
                                , pac_count as (
                                       select 
                                          YEAR
                                        , 'PAC Count' as label
                                        , count(distinct PAC_NUM) as metric_total
                                        from {orgDB}.VRDC.{table}
                                        group by 1
                                    )
                                , episode_count as (
                                       select 
                                          YEAR
                                        , 'Total Episode Count' as label
                                        , sum(NUM_EPISODES) as metric_total
                                        from {orgDB}.VRDC.{table}
                                        group by 1
                                    )
                                    select '{table}' as tableName, year, label, '{orgDB}' as orgDB, metric_total from npi_count
                                    union all 
                                    select '{table}' as tableName, year, label, '{orgDB}' as orgDB, metric_total from pac_count
                                    union all
                                    select '{table}' as tableName, year, label, '{orgDB}' as orgDB, metric_total from episode_count
                                    order by 1, 2, 3
                                    ;'''.format(table=table,orgDB=orgDB)

                sf_cursor.execute(sqlString)
                # how do you append outputs/stack them for pg and pg_network if both can have several rows?   
                df = sf_cursor.fetch_pandas_all()
                # df += df.append(df)
        return df

    # CJI_COST_SCORECARD (prod_fe)
    
    def test7(tableDict):
        print('running test 7..cji utilization in FE..')
        for table in tableDict.keys():
            if table == 'CJI_COST_SCORECARD' or table == 'CJI_COST_SCORECARD_NETWORK':
                sqlString = '''with npi_count as (
                                       select 
                                          YEAR
                                        , 'NPI Count' as label
                                        , count(distinct NPI) as metric_total
                                        from {orgDB_FE}.VRDC.{table}
                                        group by 1
                                    )
                                , pac_count as (
                                       select 
                                          YEAR
                                        , 'PAC Count' as label
                                        , count(distinct PAC_NUM) as metric_total
                                        from {orgDB_FE}.VRDC.{table}
                                        group by 1
                                    )
                                , episode_count as (
                                       select 
                                          YEAR
                                        , 'Total Episode Count' as label
                                        , sum(NUM_EPISODES) as metric_total
                                        from {orgDB_FE}.VRDC.{table}
                                        group by 1
                                    )
                                    select '{table}' as tableName, year, label, '{orgDB_FE}' as orgDB, metric_total from npi_count
                                    union all 
                                    select '{table}' as tableName, year, label, '{orgDB_FE}' as orgDB, metric_total from pac_count
                                    union all
                                    select '{table}' as tableName, year, label, '{orgDB_FE}' as orgDB, metric_total from episode_count
                                    order by 1, 2, 3
                                    ;'''.format(table=table,orgDB_FE=orgDB_FE)

                sf_cursor.execute(sqlString)
                # how do you append outputs for pg and pg_network if both can have several rows?   
                df = sf_cursor.fetch_pandas_all()
                # print(df)
                # print(dfPG)
                df7 = pd.merge(dfCJI, df, how='left', on=['TABLENAME','YEAR','LABEL'])
                df7['diff'] = df7['METRIC_TOTAL_x'] - df7['METRIC_TOTAL_y'] # does this work? 
        return df7    


    df_prod = test2(tableKeysStruct)  # has to be called
    # dupCheckTest(tableKeysStruct)    
    # print(df_prod)
    print('tables w/out pks and not included: ', tableWoutPK)

    all_test_results = {}
    output_file_path = '/Users/michael.oconnor/downloads/allTest_{orgDB}_{tableType}.xlsx'.format(orgDB=orgDB.lower(),tableType = tableType.lower())        

    writer = pd.ExcelWriter(output_file_path, engine='xlsxwriter')

    df_prod_fe = test3(tableKeysStruct)
    # print(df_prod_fe)
    # df_prod_fe.to_excel('/Users/michael.oconnor/downloads/compareTest_{orgDB}_{tableType}.xlsx'.format(orgDB=orgDB.lower(),tableType = tableType.lower()),sheet_name='compareTest')
    all_test_results['compareTest'] = df_prod_fe

    dfDups = dupCheckTest(tableKeysStruct)
    # print(dfDups)
    # dfDups.to_excel('/Users/michael.oconnor/downloads/dupTest_{orgDB}_{tableType}.xlsx'.format(orgDB=orgDB.lower(), tableType=tableType.lower()),sheet_name='dupTest')
    all_test_results['dupTest'] = dfDups

    if tableType == 'Base' or tableType == 'Network' :
        dfPG = test4(tableKeysStruct)
        # print(dfPG)
        dfPGMerge = test5(tableKeysStruct)
        # print(dfPGMerge)
        # dfPGMerge.to_excel('/Users/michael.oconnor/downloads/metricsTest_{orgDB}_{tableType}.xlsx'.format(orgDB=orgDB.lower(), tableType=tableType.lower()),sheet_name='metrics')
        all_test_results['metrics_1'] = dfPGMerge

    if tableType == 'Base' or tableType == 'Network' :
        dfCJI = test6(tableKeysStruct)
        # print(dfPG)
        dfCJIMerge = test7(tableKeysStruct)
        # print(dfPGMerge)
        # dfPGMerge.to_excel('/Users/michael.oconnor/downloads/metricsTest_{orgDB}_{tableType}.xlsx'.format(orgDB=orgDB.lower(), tableType=tableType.lower()),sheet_name='metrics')
        all_test_results['metrics_2'] = dfCJIMerge
    

    for sheet_name, df in all_test_results.items():
        df.to_excel(writer, sheet_name=sheet_name)
    writer.save()

    endTime = time.time()
    print('Time in seconds: %s' % (endTime - strtTime))
    print('Time in minutes: %s' % str(round((endTime - strtTime) / 60,2)))


except Exception as e:
    print(e)

finally:
    try:
        sf_cursor.close()
    except:
        pass
    try:
        sf_conn.close()
    except:
        pass
